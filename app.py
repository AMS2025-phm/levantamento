import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import re
import unicodedata
from flask import Flask, request, jsonify, send_from_directory # Importar send_from_directory
from flask_cors import CORS

# Configura o Flask para servir arquivos estáticos do diretório 'www'
# O 'static_url_path' define o prefixo da URL para esses arquivos.
# Deixar como '/' significa que 'www/index.html' será acessível via '/'
app = Flask(__name__, static_folder='www', static_url_path='/')

# Habilitar CORS para todas as rotas (necessário para comunicação entre domínios)
# Para produção, restrinja as origens para maior segurança. Ex:
# CORS(app, resources={r"/*": {"origins": ["capacitor://localhost", "http://localhost", "http://localhost:8100", "http://192.168.X.X", "https://seu-frontend-render.onrender.com"]}})
CORS(app)

# Nome do arquivo onde os dados das localidades são armazenados (se o backend ainda os usa)
ARQUIVO_LOCALIDADES_JSON = "localidades.json"

# --- Configurações de E-mail (Lidas de variáveis de ambiente do Render) ---
EMAIL_USER = os.environ.get("EMAIL_USER")
EMAIL_PASSWORD = os.environ.get("EMAIL_PASSWORD")
SMTP_SERVER = os.environ.get("SMTP_SERVER")
SMTP_PORT = int(os.environ.get("SMTP_PORT", 587))

# --- Funções Auxiliares ---
def carregar_dados_localidades():
    if os.path.exists(ARQUIVO_LOCALIDADES_JSON):
        try:
            with open(ARQUIVO_LOCALIDADES_JSON, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            print(f"ERRO: Falha ao decodificar {ARQUIVO_LOCALIDADES_JSON}: {e}")
            return {}
    print(f"AVISO: {ARQUIVO_LOCALIDADES_JSON} não encontrado no backend. As rotas relacionadas podem falhar.")
    return {}

def slugify(value, allow_unicode=False):
    value = str(value)
    if allow_unicode:
        value = unicodedata.normalize('NFKC', value)
    else:
        value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^\w\s-]', '', value).strip().lower()
    return re.sub(r'[-\s]+', '-', value)

# --- Rotas para Servir Arquivos Estáticos do Frontend ---
@app.route('/')
def serve_index():
    """Serve o index.html na raiz do aplicativo."""
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    """Serve outros arquivos estáticos (CSS, JS, imagens) do diretório www."""
    # Garante que o service worker seja servido com o cabeçalho correto
    if path == 'sw.js':
        return send_from_directory(app.static_folder, path, mimetype='application/javascript')
    return send_from_directory(app.static_folder, path)

# --- Suas Rotas Existentes para Seleção de Local/Ambiente/Descrição (se ainda forem usadas pelo backend) ---
@app.route('/get_localidades_unidades', methods=['GET'])
def get_localidades_unidades():
    """Retorna a lista de localidades e unidades para o dropdown em JSON."""
    localidades = carregar_dados_localidades()
    lista_localidades_unidades = []
    for local, unidades in localidades.items():
        for unidade in unidades.keys():
            lista_localidades_unidades.append(f"{local} - {unidade}")

    lista_localidades_unidades.sort()
    return jsonify(lista_localidades_unidades), 200

@app.route('/get_unidade_data/<string:local_unidade>', methods=['GET'])
def get_unidade_data(local_unidade):
    """Retorna os detalhes de uma unidade específica."""
    if " - " not in local_unidade:
        return jsonify({"status": "error", "message": "Formato de unidade inválido."}), 400

    local, unidade = local_unidade.split(" - ", 1)
    localidades = carregar_dados_localidades()

    if local in localidades and unidade in localidades[local]:
        return jsonify({"status": "success", "data": localidades[local][unidade]}), 200
    else:
        return jsonify({"status": "error", "message": "Unidade não encontrada."}), 404

# --- Rota para RECEBER DADOS DO FORMULÁRIO DO FRONTEND e processar ---
@app.route('/submit_levantamento', methods=['POST'])
def submit_levantamento():
    try:
        dados_formulario = request.get_json()
        if not dados_formulario:
            return jsonify({"status": "error", "message": "Nenhum dado de formulário recebido."}), 400

        print(f"Dados de levantamento recebidos para processamento: {json.dumps(dados_formulario, indent=2)}")

        localidade = dados_formulario.get('localidade', 'N/A')
        unidade = dados_formulario.get('unidade', 'N/A')
        data_coleta = dados_formulario.get('dataColeta', datetime.date.today().strftime('%Y-%m-%d'))
        responsavel = dados_formulario.get('responsavel', 'N/A')
        contato_email = dados_formulario.get('contatoEmail', '')

        medidas = dados_formulario.get('medidas', [])

        # --- Geração do XLSX ---
        output = io.BytesIO()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Levantamento de Medidas"

        headers = [
            "Localidade", "Unidade", "Data da Coleta", "Responsável", "Email de Contato",
            "Ambiente", "Tipo de Medida", "Descrição", "Medida L", "Medida C", "Quantidade",
            "Detalhes Adicionais", "Tipo de Piso", "Tipo de Parede", "Observações"
        ]
        sheet.append(headers)

        for medida in medidas:
            row = [
                localidade, unidade, data_coleta, responsavel, contato_email,
                medida.get('ambiente', ''),
                medida.get('tipoMedida', ''),
                medida.get('descricao', ''),
                medida.get('medidaL', ''),
                medida.get('medidaC', ''),
                medida.get('quantidade', ''),
                medida.get('detalhesAdicionais', ''),
                medida.get('tipoPiso', ''),
                medida.get('tipoParede', ''),
                medida.get('observacoes', '')
            ]
            sheet.append(row)

        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        workbook.save(output)
        output.seek(0)

        # --- Envio de E-mail com o XLSX Anexado ---
        if not all([EMAIL_USER, EMAIL_PASSWORD, SMTP_SERVER, SMTP_PORT]):
            print("AVISO: Credenciais de e-mail incompletas. E-mail não será enviado.")
            # return jsonify({"status": "error", "message": "Configurações de e-mail incompletas no servidor."}), 500

        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = contato_email
        msg['Subject'] = f"Levantamento de Medidas - {localidade} - {unidade} ({data_coleta})"

        msg.attach(MIMEText("Prezado(a),\n\nSegue em anexo o levantamento de medidas realizado.\n\nAtenciosamente,\nSua Equipe", 'plain'))

        part = MIMEBase('application', 'octet-stream')
        part.set_payload(output.read())
        encoders.encode_base64(part)
        filename = f"Levantamento_Medidas_{slugify(localidade)}_{slugify(unidade)}_{data_coleta}.xlsx"
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)

        try:
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
                smtp.starttls()
                smtp.login(EMAIL_USER, EMAIL_PASSWORD)
                smtp.send_message(msg)
            print(f"E-mail com XLSX enviado para {contato_email} com sucesso!")
        except Exception as e:
            print(f"ERRO ao enviar e-mail com XLSX para {contato_email}: {e}")

        return jsonify({"status": "success", "message": "Dados recebidos, XLSX gerado e e-mail enviado!", "data_received": dados_formulario}), 200

    except Exception as e:
        print(f"ERRO CRÍTICO no submit_levantamento: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": "Erro interno do servidor ao processar levantamento", "details": str(e)}), 500

# Endpoint de saúde para o Render verificar se o app está ativo
@app.route('/healthcheck', methods=['HEAD', 'GET'])
def healthcheck():
    return '', 200

if __name__ == '__main__':
    # Quando rodando localmente, configure as variáveis de ambiente manualmente ou via .env
    # No Render, elas serão injetadas automaticamente.
    # Exemplo para teste local (REMOVA OU COMENTE EM PRODUÇÃO):
    # os.environ['EMAIL_USER'] = 'seu_email@exemplo.com'
    # os.environ['EMAIL_PASSWORD'] = 'sua_senha_ou_senha_de_app'
    # os.environ['SMTP_SERVER'] = 'smtp.seuservidor.com' # Ex: 'smtp.gmail.com'
    # os.environ['SMTP_PORT'] = '587' # Ou 465 para SSL

    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=True)